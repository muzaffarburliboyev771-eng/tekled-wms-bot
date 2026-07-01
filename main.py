
import sqlite3
import asyncio
import os
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from openpyxl import load_workbook

TOKEN = "8953739174:AAHG2U4icxh_xs-8Z2v_tH4ey_40T8qI5DI"

ALLOWED_USERS = [
    5528895597, 
    7489557412,
]

bot = Bot(token=TOKEN)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)


class EditProduct(StatesGroup):
    waiting_for_field = State()
    waiting_for_new_value = State()
    current_code = None
    field_to_edit = None

def get_db():
    return sqlite3.connect("wms.db")

@dp.message(Command("start"))
async def start(msg: types.Message):
    if msg.from_user.id not in ALLOWED_USERS:
        await msg.answer("❌ Sizga ruxsat yo'q")
        return

    await msg.answer(
        "📦 WMS Bot\n\n"
        "➕ Qo'shish: Mahsulot | Joy | Kod\n"
        "🔍 Qidirish: A312\n"
        "📤 Excel yuklash: .xlsx fayl yuboring"
    )

@dp.message(lambda m: m.document)
async def excel_upload(msg: types.Message):
    if msg.from_user.id not in ALLOWED_USERS:
        return

    if not msg.document.file_name.endswith(".xlsx"):
        await msg.answer("❌ Faqat .xlsx fayl yuboring")
        return

    file = await bot.get_file(msg.document.file_id)
    file_path = f"temp_{msg.document.file_name}"
    await bot.download_file(file.file_path, file_path)

    conn = get_db()
    cur = conn.cursor()
    wb = load_workbook(file_path)
    ws = wb.active

    added, skipped = 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            name, location, code = row
            if not name or not code:
                continue
            cur.execute(
                "INSERT INTO products (product_name, location, location_code) VALUES (?, ?, ?)",
                (str(name), str(location), str(code).upper())
            )
            added += 1
        except:
            skipped += 1

    conn.commit()
    conn.close()
    os.remove(file_path)

    await msg.answer(f"✅ Yuklandi\n📥 Qo'shildi: {added}\n⚠️ O'tkazildi: {skipped}")

@dp.callback_query(F.data.startswith("edit_"))
async def edit_product_callback(query: types.CallbackQuery, state: FSMContext):
    code = query.data.split("_")[1]
    field = query.data.split("_")[2]
    
    await query.answer()
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT product_name, location, location_code FROM products WHERE location_code=?", (code,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        await query.message.answer("❌ Mahsulot topilmadi")
        return
    
    if field == "location":
        current_value = row[1]
        field_name = "Joyi"
    else:  # code
        current_value = row[2]
        field_name = "Kodi"
    
    # Tasdiqlovchi tugmalar
    confirmation_keyboard = types.InlineKeyboardMarkup(
        inline_keyboard=[[
            types.InlineKeyboardButton(text="✅ Ha, tahrirlash", callback_data=f"confirm_edit_{code}_{field}"),
            types.InlineKeyboardButton(text="❌ Yo'q", callback_data="cancel_edit")
        ]]
    )
    
    await query.message.answer(
        f"🔄 <b>{field_name}</b> ni tahrirlashni tasdiqlaysizmi?\n"
        f"Hozirgi qiymati: <b>{current_value}</b>",
        reply_markup=confirmation_keyboard,
        parse_mode="HTML"
    )

@dp.callback_query(F.data.startswith("confirm_edit_"))
async def confirm_edit(query: types.CallbackQuery, state: FSMContext):
    parts = query.data.split("_")
    code = parts[2]
    field = parts[3]
    
    await query.answer()
    
    if field == "location":
        field_name = "Joyi"
    else:
        field_name = "Kodi"
    
    await state.set_state(EditProduct.waiting_for_new_value)
    await state.update_data(current_code=code, field_to_edit=field)
    
    await query.message.answer(
        f"🔧 {field_name} ni tahrirlang\n"
        f"Yangi qiymatni yozing:",
        parse_mode="HTML"
    )

@dp.callback_query(F.data == "cancel_edit")
async def cancel_edit(query: types.CallbackQuery):
    await query.answer("❌ Tahrirlash bekor qilindi")
    await query.message.delete()

@dp.message(EditProduct.waiting_for_new_value)
async def handle_new_value(msg: types.Message, state: FSMContext):
    if msg.from_user.id not in ALLOWED_USERS:
        return
    
    data = await state.get_data()
    code = data['current_code']
    field = data['field_to_edit']
    new_value = msg.text.strip()
    
    if not new_value:
        await msg.answer("❌ Qiymat bo'sh bo'la olmaydi")
        return
    
    conn = get_db()
    cur = conn.cursor()
    
    try:
        if field == "location":
            cur.execute(
                "UPDATE products SET location=? WHERE location_code=?",
                (new_value, code)
            )
            field_name = "Joyi"
        else:  # code
            new_value = new_value.upper()
            cur.execute(
                "UPDATE products SET location_code=? WHERE location_code=?",
                (new_value, code)
            )
            field_name = "Kodi"
        
        conn.commit()
        await msg.answer(f"✅ {field_name} muvaffaqiyatli tahrirlandi!")
        
    except sqlite3.IntegrityError:
        await msg.answer("❌ Bu kod allaqachon mavjud")
    except Exception as e:
        await msg.answer(f"❌ Xatolik: {str(e)}")
    finally:
        conn.close()
        await state.clear()

@dp.message()
async def handler(msg: types.Message):
    if msg.from_user.id not in ALLOWED_USERS:
        return

    text = msg.text.strip()
    conn = get_db()
    cur = conn.cursor()

    if "|" in text:
        try:
            name, location, code = [x.strip() for x in text.split("|")]
            cur.execute(
                "INSERT INTO products (product_name, location, location_code) VALUES (?, ?, ?)",
                (name, location, code.upper())
            )
            conn.commit()
            await msg.answer("✅ Saqlandi")
        except:
            await msg.answer("❌ Xatolik (kod mavjud bo'lishi mumkin)")
    elif text.lower().startswith("edit "):
        # "edit KOD" buyrugi
        code = text.split(" ", 1)[1].strip().upper()
        cur.execute(
            "SELECT product_name, location, location_code FROM products WHERE location_code=?",
            (code,)
        )
        row = cur.fetchone()
        
        if row:
            keyboard = types.InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        types.InlineKeyboardButton(
                            text="✏️ Joyi tahrirlash",
                            callback_data=f"edit_{row[2]}_location"
                        ),
                        types.InlineKeyboardButton(
                            text="✏️ Kodi tahrirlash",
                            callback_data=f"edit_{row[2]}_code"
                        )
                    ]
                ]
            )
            await msg.answer(
                f"📦 Mahsulot: {row[0]}\n📍 Joyi: <b>{row[1]}</b>\n🔖 Kodi: {row[2]}",
                reply_markup=keyboard,
                parse_mode="HTML"
            )
        else:
            await msg.answer("❌ Mahsulot topilmadi")
    else:
        # Oddiy qidirish - tugmalarsiz
        cur.execute(
            "SELECT product_name, location, location_code FROM products WHERE location_code=?",
            (text.upper(),)
        )
        row = cur.fetchone()
        if row:
            await msg.answer(
                f"📦 Mahsulot: {row[0]}\n📍 Joyi: <b>{row[1]}</b>\n🔖 Kodi: {row[2]}",
                parse_mode="HTML"
            )
        else:
            await msg.answer("❌ Topilmadi")

    conn.close()

async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
